"""
売上明細 vs 請求CSV 照合チェックツール（全校舎対応）

config.yaml に校舎を追加するだけで、新校舎の照合チェックに対応。
売上申請（Excel）に記載された金額が、請求データ（CSV）と一致しているか検証する。
"""

import csv
import re
import sys
import os
import glob
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime

try:
    import yaml
except ImportError:
    print("PyYAML が必要です: pip install pyyaml")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)


# ====================================================================
# データクラス
# ====================================================================

@dataclass(frozen=True)
class SchoolConfig:
    name: str
    excel_dir: str
    excel_pattern: str
    sheet_index: int = 2
    school_keywords: tuple[str, ...] = ()  # クラス情報の校舎名マッチ用


@dataclass(frozen=True)
class MonthPair:
    year: int
    month: int
    csv_path: str
    excel_path: str

    @property
    def label(self) -> str:
        return f"{self.year}年{self.month}月"


@dataclass
class CheckResult:
    school: str
    month_label: str
    result_type: str      # "TOTAL_DIFF", "TOTAL_MATCH", "NOT_IN_CSV"
    sid: str
    name: str
    row: int
    diffs: list = field(default_factory=list)
    excel_total: float = 0.0
    csv_total: float = 0.0
    # 前後月の引落額: {月ラベル: {列レター: 金額}}
    monthly_billing: dict = field(default_factory=dict)
    # 備考: {列レター: コメント文字列}
    remarks: dict = field(default_factory=dict)
    grade: str = ""
    # 対象月の前後月ラベルリスト（表示順）
    month_columns: list = field(default_factory=list)


# ====================================================================
# ブランド名 → Excel列 のマッピング
# Excel列構造:
#   D=設備費, E=授業料(学習塾), F=月会費(学習塾), G=テスト対策/講習会費, H=模試代, I=講習会費
#   J=授業料(プログラミング), K=月会費(プログラミング), L=講習会費(プログラミング)
#   M=授業料(アン), N=月会費(アン)
#   O=授業料(そろばん), P=月会費(そろばん), Q=講習会費(そろばん)
#   R=授業料(筆っこ), S=月会費(筆っこ), T=講習会費(筆っこ)
#   U=授業料(将棋), V=月会費(将棋), W=講習会費(将棋)
#   X=入会金, Y=その他, Z=売上合計
# ====================================================================


BRAND_COLUMN_MAP = {
    "アンイングリッシュクラブ": ("M", "N"),
    "アンそろばんクラブ": ("O", "P"),
    "アンそろばんクラブ【選択講習会】": ("O", "P"),
    "アンプログラミングクラブ": ("J", "K"),
    "アン美文字クラブ": ("R", "S"),
    "アン将棋クラブ": ("U", "V"),
}

# 既定値（小幡レイアウトをバックアップとして保存）
_DEFAULT_BRAND_COLUMN_MAP = dict(BRAND_COLUMN_MAP)

# シート内のグループ見出し名 → CSVブランド名 の対応。
# 校舎ごとに異なる呼称を吸収する（筆っこ/美文字 など同一ブランドの別名もここで統一）。
GROUP_LABEL_TO_CSV_BRAND = {
    "プログラミング": "アンプログラミングクラブ",
    "アン": "アンイングリッシュクラブ",
    "そろばん": "アンそろばんクラブ",
    "筆っこ": "アン美文字クラブ",
    "美文字": "アン美文字クラブ",
    "将棋": "アン将棋クラブ",
    # 学童 は CSVブランドに対応がないため学習塾扱い
}

COL_DISPLAY = {
    "D": "設備費",
    "E": "授業料(学習塾)",
    "F": "月会費(学習塾)",
    "G": "テスト対策/講習会費",
    "H": "模試代",
    "I": "講習会費",
    "J": "授業料(プログラミング)",
    "K": "月会費(プログラミング)",
    "L": "講習会費(プログラミング)",
    "M": "授業料(アン)",
    "N": "月会費(アン)",
    "O": "授業料(そろばん)",
    "P": "月会費(そろばん)",
    "Q": "講習会費(そろばん)",
    "R": "授業料(筆っこ)",
    "S": "月会費(筆っこ)",
    "T": "講習会費(筆っこ)",
    "U": "授業料(将棋)",
    "V": "月会費(将棋)",
    "W": "講習会費(将棋)",
    "X": "入会金",
    "Y": "その他",
    "Z": "売上合計",
}


# ====================================================================
# 設定読み込み
# ====================================================================

def load_config(config_path: str = None) -> dict:
    if config_path is None:
        config_path = str(Path(__file__).parent / "config.yaml")
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


# ====================================================================
# ファイル自動検出
# ====================================================================

def discover_csv_files(csv_base_dir: str) -> dict[tuple[int, int], str]:
    """
    請求CSVを自動検出。ディレクトリ名から請求対象月を推定。
    CSVディレクトリの日付 +1ヶ月 = 請求対象月
    返り値: {(year, month): csv_path}
    """
    result = {}
    base = Path(csv_base_dir)

    for d in sorted(base.iterdir()):
        if not d.is_dir():
            continue
        # ディレクトリ名から日付を抽出 (YYYYMMDD_...)
        m = re.match(r"(\d{4})(\d{2})\d{2}", d.name)
        if not m:
            continue

        # CSVファイルを探す
        csvs = list(d.glob("AC_5_*UTF8*.csv"))
        if not csvs:
            continue

        dir_year = int(m.group(1))
        dir_month = int(m.group(2))

        # 請求対象月 = ディレクトリの月 + 1
        billing_month = dir_month + 1
        billing_year = dir_year
        if billing_month > 12:
            billing_month = 1
            billing_year += 1

        # 最新のCSVを使用（同じディレクトリに複数ある場合）
        csv_path = str(sorted(csvs, key=lambda p: p.stat().st_mtime)[-1])
        result[(billing_year, billing_month)] = csv_path

    return result


def discover_excel_files(school: SchoolConfig) -> dict[tuple[int, int], str]:
    """
    校舎のExcelファイルを自動検出。ファイル名から年月を抽出。
    返り値: {(year, month): excel_path}
    """
    result = {}
    excel_dir = Path(school.excel_dir)

    if not excel_dir.exists():
        return result

    for f in excel_dir.iterdir():
        if not f.is_file():
            continue
        if not (f.suffix == ".xlsm" or f.suffix == ".xlsx"):
            continue

        # ファイル名から (YYYY.MM) を抽出
        m = re.search(r"\((\d{4})\.(\d{2})\)", f.name)
        if not m:
            continue

        year = int(m.group(1))
        month = int(m.group(2))
        result[(year, month)] = str(f)

    return result


def match_months(
    csv_files: dict[tuple[int, int], str],
    excel_files: dict[tuple[int, int], str],
) -> list[MonthPair]:
    """CSV と Excel を月でマッチング"""
    pairs = []
    for ym in sorted(set(csv_files.keys()) & set(excel_files.keys())):
        pairs.append(MonthPair(
            year=ym[0],
            month=ym[1],
            csv_path=csv_files[ym],
            excel_path=excel_files[ym],
        ))
    return pairs


# ====================================================================
# ユーザークラス情報（校舎別受講フィルタ）
# ====================================================================

def discover_class_info_files(
    class_info_dir: str,
) -> dict[tuple[int, int], str]:
    """
    クラス情報Excelを自動検出。ファイル名から月を抽出。
    返り値: {(year, month): path}
    """
    result = {}
    d = Path(class_info_dir)
    if not d.exists():
        return result

    month_map = {
        "1月": 1, "2月": 2, "3月": 3, "4月": 4, "5月": 5, "6月": 6,
        "7月": 7, "8月": 8, "9月": 9, "10月": 10, "11月": 11, "12月": 12,
    }

    for f in d.iterdir():
        if not f.is_file() or f.suffix not in (".xlsx", ".xlsm"):
            continue
        # "1月末.xlsx", "11月分.xlsx" etc.
        m = re.search(r"(\d{1,2})月", f.name)
        if not m:
            continue
        month = int(m.group(1))
        # 年は修正日時から推定（ファイル名に年がないため）
        year = datetime.fromtimestamp(f.stat().st_mtime).year
        result[(year, month)] = str(f)

    return result


def read_class_info(
    class_info_path: str,
    school_keywords: tuple[str, ...],
) -> dict[str, set[str]]:
    """
    クラス情報を読み込み、指定校舎に在籍する生徒のブランド一覧を返す。
    返り値: {生徒ID: {この校舎で受講しているブランド名のset}}

    校舎名にschool_keywordsのいずれかが含まれていれば、その校舎のブランドとみなす。
    """
    wb = openpyxl.load_workbook(class_info_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    sid_brands: dict[str, set[str]] = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        sid_val = row[6].value   # G=生徒ID
        brand = row[25].value    # Z=ブランド名
        school = row[31].value   # AF=校舎名

        if sid_val is None or school is None:
            continue

        sid = str(int(sid_val)) if isinstance(sid_val, (int, float)) else str(sid_val)
        school_str = str(school)

        # この校舎に該当するか
        if any(kw in school_str for kw in school_keywords):
            if sid not in sid_brands:
                sid_brands[sid] = set()
            if brand:
                sid_brands[sid].add(str(brand))

    wb.close()
    return sid_brands


def read_all_class_sids(
    class_info_path: str,
    school_keywords: tuple[str, ...],
) -> set[str]:
    """クラス情報に存在する校舎の全生徒IDを返す"""
    wb = openpyxl.load_workbook(class_info_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sids = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        sid_val = row[6].value
        school = row[31].value
        if sid_val is None or school is None:
            continue
        if any(kw in str(school) for kw in school_keywords):
            sids.add(str(int(sid_val)) if isinstance(sid_val, (int, float)) else str(sid_val))
    wb.close()
    return sids


def read_mid_month_withdrawals(
    class_info_path: str,
    school_keywords: tuple[str, ...],
    target_year: int,
    target_month: int,
) -> dict[str, dict[str, str]]:
    """
    当月中または前月末に退会日があるブランドをブランド単位で検出。
    返り値: {生徒ID: {ブランド名: "退会日"}}
    """
    from datetime import date
    target_first = date(target_year, target_month, 1)
    if target_month == 12:
        next_first = date(target_year + 1, 1, 1)
    else:
        next_first = date(target_year, target_month + 1, 1)

    wb = openpyxl.load_workbook(class_info_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result: dict[str, dict[str, str]] = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        sid_val = row[6].value
        school = row[31].value
        brand = row[25].value
        taikai = row[21].value
        if sid_val is None or school is None or taikai is None:
            continue
        if not any(kw in str(school) for kw in school_keywords):
            continue

        sid = str(int(sid_val)) if isinstance(sid_val, (int, float)) else str(sid_val)
        brand_str = str(brand) if brand else ""

        try:
            if hasattr(taikai, "date"):
                td = taikai.date()
            elif hasattr(taikai, "year"):
                td = taikai
            else:
                td = datetime.strptime(str(taikai).split()[0], "%Y-%m-%d").date()
            # 退会日が前月末～当月内
            from datetime import timedelta
            prev_last = target_first - timedelta(days=1)
            if td >= prev_last and td < next_first:
                if sid not in result:
                    result[sid] = {}
                if brand_str:
                    result[sid][brand_str] = str(td)
        except (ValueError, AttributeError):
            pass

    wb.close()
    return result


def read_withdrawn_students(
    class_info_path: str,
    school_keywords: tuple[str, ...],
    target_year: int,
    target_month: int,
) -> set[str]:
    """
    対象月の初日時点で既に退会済みの生徒IDを返す。
    退会日 < 対象月の初日 → 退会者とみなす。
    """
    from datetime import date

    target_first = date(target_year, target_month, 1)

    wb = openpyxl.load_workbook(class_info_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 生徒ごとに全校舎・全ブランドの退会日を集める
    # この校舎のクラスが全て退会済みなら退会者とみなす
    sid_classes: dict[str, list[tuple[bool, bool]]] = {}  # {sid: [(is_this_school, is_withdrawn)]}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        sid_val = row[6].value   # G=生徒ID
        school = row[31].value   # AF=校舎名
        taikai = row[21].value   # V=退会日

        if sid_val is None or school is None:
            continue

        sid = str(int(sid_val)) if isinstance(sid_val, (int, float)) else str(sid_val)
        school_str = str(school)

        is_this_school = any(kw in school_str for kw in school_keywords)
        if not is_this_school:
            continue

        # 退会日の判定
        is_withdrawn = False
        if taikai is not None:
            try:
                if hasattr(taikai, "date"):
                    taikai_date = taikai.date()
                elif hasattr(taikai, "year"):
                    taikai_date = taikai
                else:
                    taikai_date = datetime.strptime(
                        str(taikai).split()[0], "%Y-%m-%d"
                    ).date()
                is_withdrawn = taikai_date < target_first
            except (ValueError, AttributeError):
                pass

        if sid not in sid_classes:
            sid_classes[sid] = []
        sid_classes[sid].append(is_withdrawn)

    wb.close()

    # この校舎の全クラスが退会済みの生徒
    withdrawn = set()
    for sid, classes in sid_classes.items():
        if classes and all(classes):
            withdrawn.add(sid)

    return withdrawn


def filter_billing_by_school(
    billing_entries: list[tuple[str, str, float]],
    school_brands: set[str] | None,
) -> list[tuple[str, str, float]]:
    """
    請求明細から、この校舎で受講していないブランドの項目を除外する。
    school_brands が None の場合はフィルタなし（全項目を返す）。

    設備費・入会金・家族割等のブランド横断項目はフィルタしない。
    """
    if school_brands is None:
        return billing_entries

    # フィルタ不要のカテゴリ（ブランドに依存しない費目）
    PASSTHROUGH_CATEGORIES = {
        "設備費", "入会金", "模試代", "家族割", "割引", "過不足金",
        "諸経費", "家賃", "総合指導管理費", "0", "1",
        "入会時教材費", "入会時授業料1", "入会時授業料2",
        "入会時授業料3", "入会時授業料A", "入会時月会費", "入会時設備費",
    }

    # BRAND_COLUMN_MAP にあるブランドのみ校舎フィルタ対象
    # それ以外（学習塾系）はクラス情報に載っていなくても通す
    FILTERED_BRANDS = set(BRAND_COLUMN_MAP.keys())

    filtered = []
    for brand, category, amount in billing_entries:
        # ブランド横断の費目はそのまま通す
        if category in PASSTHROUGH_CATEGORIES:
            filtered.append((brand, category, amount))
            continue
        # ブランドが空の場合（社割等）はそのまま通す
        if not brand:
            filtered.append((brand, category, amount))
            continue
        # BRAND_COLUMN_MAP に含まれるブランド → 校舎フィルタで判定
        if brand in FILTERED_BRANDS:
            if brand in school_brands:
                filtered.append((brand, category, amount))
            elif any(sb in brand or brand in sb for sb in school_brands):
                filtered.append((brand, category, amount))
        else:
            # 学習塾系ブランド → クラス情報になくても通す
            filtered.append((brand, category, amount))

    return filtered


# ====================================================================
# データ読み込み・集計（コアロジック）
# ====================================================================

def parse_number(val) -> float:
    if val is None or val == "" or val == 0:
        return 0.0
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def read_billing_csv(csv_path: str) -> dict[str, list[tuple[str, str, float]]]:
    """請求CSVを読み込み → {生徒ID: [(ブランド名, カテゴリ名, 月額料金), ...]}"""
    result = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sid = row["生徒ID"].strip()
            brand = row["ブランド名"].strip()
            category = row["請求カテゴリ名"].strip()
            amount = parse_number(row["月額料金"])
            if sid not in result:
                result[sid] = []
            result[sid].append((brand, category, amount))
    return result


# 学年情報キャッシュ
_grade_cache: dict[str, dict[str, str]] = {}


def read_grades_from_csv(csv_path: str) -> dict[str, str]:
    """請求CSVから生徒IDと学年の対応を取得 → {生徒ID: 学年}"""
    if csv_path in _grade_cache:
        return _grade_cache[csv_path]
    grades = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sid = row["生徒ID"].strip()
            grade = row.get("学年", "").strip()
            if sid and grade and sid not in grades:
                grades[sid] = grade
    _grade_cache[csv_path] = grades
    return grades


def aggregate_csv_for_student(
    entries: list[tuple[str, str, float]],
) -> dict[str, float]:
    """CSV明細をExcel列に対応する形で集計 → {列レター: 金額}"""
    col_totals: dict[str, float] = {}

    for brand, category, amount in entries:
        if amount == 0:
            continue

        target_col = _map_to_column(brand, category)

        if target_col:
            col_totals[target_col] = col_totals.get(target_col, 0) + amount

    return col_totals


def _map_to_column(brand: str, category: str) -> str | None:
    """ブランド名＋カテゴリ名 → Excel列レターを決定"""
    if category == "設備費":
        return "D"
    if category == "入会金":
        return "X"
    if category == "模試代":
        return "H"
    if category in ("0", "1"):
        # 「その他/0」「その他/1」は講習会費やテスト対策費の可能性あり
        # → Y列にマッピングするが、アラートで別途警告
        return "Y"
    if category in ("入会時教材費", "入会時授業料1", "入会時授業料2",
                     "入会時授業料3", "入会時授業料A", "入会時月会費",
                     "入会時設備費"):
        return "Y"
    if category in ("家族割", "割引", "過不足金", "諸経費", "家���",
                     "総合指導管理費"):
        return "Y"

    if brand in BRAND_COLUMN_MAP:
        jugyou_col, getsukai_col = BRAND_COLUMN_MAP[brand]
        if category in ("授業料", "追加授業料"):
            return jugyou_col
        if category == "月会費":
            return getsukai_col
        if category in ("講習会費", "必須講座", "必須講習会", "テスト対策"):
            if brand.startswith("アンそろばん"):
                return "Q"
            if brand == "アンプログラミングクラブ":
                return "L"
            if brand == "アン美文字クラブ":
                return "T"
            if brand == "アン将棋クラブ":
                return "W"
            return "I"
        # カテゴリ空欄でブランド名に「講習会」「検定」が含まれる場合
        if not category or category == "":
            if "講習会" in brand:
                if brand.startswith("アンそろばん"):
                    return "Q"
                if "美文字" in brand:
                    return "T"
                if "将棋" in brand:
                    return "W"
                if "プログラミング" in brand:
                    return "L"
                return "I"
            if "検定" in brand:
                return "Y"
        return "Y"

    # 上記BRAND_COLUMN_MAP以外のブランドは全て学習塾として扱う
    if category in ("授業料", "追加授業料"):
        return "E"
    if category == "月会費":
        return "F"
    if category in ("講習会費", "必須講座", "必須講習会", "テスト対策"):
        return "I"
    return "Y"


def _find_header_row(ws) -> int | None:
    """ヘッダ行(生徒ID を含む最初の行)を検出。1〜8行目を探索。"""
    for r in range(1, 9):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if v and "生徒ID" in str(v):
                return r
    return None


def _score_sales_sheet(ws) -> int:
    """
    売上明細シートらしさを採点。要件:
    1) 生徒ID を含むヘッダ行がある
    2) ヘッダ行に 授業料/月会費/講習会費/月謝 等の金額列キーワードがある
    3) ヘッダ行+1 以降で B列=数値ID & C列=氏名 の行を数える
    金額キーワードが無ければ 0 (生徒マスターシート等を除外)。
    """
    header_row = _find_header_row(ws)
    if header_row is None:
        return 0

    money_keywords = ("授業料", "月会費", "講習会費", "月謝", "講習代", "模試代", "入会金")
    has_money = False
    for c in range(1, 40):
        v = ws.cell(row=header_row, column=c).value
        if v:
            vstr = str(v).replace("\n", "").replace(" ", "")
            if any(kw in vstr for kw in money_keywords):
                has_money = True
                break
    if not has_money:
        return 0

    score = 0
    max_scan = min(ws.max_row, 500)
    for r in range(header_row + 1, max_scan + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        c_val = ws.cell(row=r, column=3).value
        if a_val is not None and isinstance(a_val, str) and "計" in a_val:
            break
        if isinstance(b_val, (int, float)) and isinstance(c_val, str) and c_val.strip():
            score += 1
    return score


def detect_sales_sheet_index(excel_path: str) -> int | None:
    """ブック内から売上明細シートを自動検出。5人未満しか見つからなければ None。"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        best_idx, best_score = None, 0
        for i, name in enumerate(wb.sheetnames):
            s = _score_sales_sheet(wb[name])
            if s > best_score:
                best_idx, best_score = i, s
        return best_idx if best_score >= 5 else None
    finally:
        wb.close()


def _normalize_label(s) -> str:
    if s is None:
        return ""
    return str(s).replace("\u3000", "").replace(" ", "").replace("\n", "").strip()


# タイプB (丸一/北山/金屋) 用: 列名から直接 (brand, item) を取り出す
_TYPE_B_ITEMS = ("授業料", "月会費", "講習会費", "テスト対策費", "模試代")
_TYPE_B_BRAND_NORMALIZE = {
    "塾": "学習塾",
    "学習塾": "学習塾",
    "プログラミング": "プログラミング",
    "プロ": "プログラミング",
    "アン": "アン",
    "そろばん": "そろばん",
    "筆っこ": "筆っこ",
    "美文字": "美文字",
    "将棋": "将棋",
    "学童": "学童",
}


def _parse_type_b_label(clean: str):
    """'将棋授業料','プログラミング月会費','塾月会費','講習会費テスト対策費' などを分解。
    返り値: (brand, item) または None。brand が空なら item 単独扱いとして _leading/_trailing に分類。"""
    if not clean:
        return None
    # 複合: "講習会費テスト対策費" → 学習塾の テスト対策講習会費
    if "テスト対策" in clean and "講習会費" in clean:
        return ("学習塾", "テスト対策講習会費")
    for item in _TYPE_B_ITEMS:
        if clean.endswith(item):
            brand_part = clean[:-len(item)]
            if not brand_part:
                return ("", item)  # ブランド前置なし (タイプC の単独列)
            brand_norm = _TYPE_B_BRAND_NORMALIZE.get(brand_part, brand_part)
            return (brand_norm, item)
    return None


def _parse_type_b_header(item_labels: dict, header_row: int) -> dict:
    """単一行ヘッダ形式(タイプB)を解析してレイアウト辞書を返す。
    brand+item 埋め込みラベルが2つ以上見つかれば有効。"""
    from openpyxl.utils import get_column_letter

    structured = {}
    col_display = {}
    STANDALONE_LEADING = {"設備費", "空調費"}
    STANDALONE_TRAILING = {"入会金", "その他", "その他割引", "成績優秀割引", "売上", "売上合計", "備考"}
    SKIP = {"No.", "生徒ID", "生徒氏名", "Name", "学年", "コース①", "コース②", "備考"}

    brand_hits = 0
    for c, lbl in item_labels.items():
        clean = lbl.replace("\n", "").replace(" ", "").replace("\u3000", "")
        if clean in SKIP:
            continue
        letter = get_column_letter(c)
        if clean in STANDALONE_LEADING:
            structured[letter] = ("_leading", clean)
            col_display[letter] = clean
            continue
        if clean in STANDALONE_TRAILING:
            structured[letter] = ("_trailing", clean)
            col_display[letter] = clean
            continue
        if clean == "模試代":
            structured[letter] = ("学習塾", "模試代")
            col_display[letter] = "模試代(学習塾)"
            continue
        if clean == "講習会費":
            # タイプBでは単独の講習会費は学習塾扱い
            structured[letter] = ("学習塾", "講習会費")
            col_display[letter] = "講習会費(学習塾)"
            continue
        parsed = _parse_type_b_label(clean)
        if parsed:
            brand, item = parsed
            if brand:
                structured[letter] = (brand, item)
                col_display[letter] = f"{item}({brand})"
                brand_hits += 1
            else:
                structured[letter] = ("_trailing", item)
                col_display[letter] = item

    if brand_hits < 2:
        return {}  # タイプB と認められない

    return {
        "col_display": col_display,
        "structured": structured,
        "header_row": header_row,
    }


def detect_column_layout(excel_path: str, sheet_index: int) -> dict:
    """
    売上明細のヘッダ行（3行目=ブランド見出し, 4行目=項目名）を解析し、
    校舎ごとの列マッピングを動的に構築する。

    戻り値: {
        "col_display": {letter: "授業料(アン)" ...},
        "brand_column_map": {"アンイングリッシュクラブ": ("L","M"), ...},
        "koshukai_cols": set(),
        "recurring_cols": set(),
    }
    検出失敗時は空dict。
    """
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
    try:
        if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
            return {}
        ws = wb[wb.sheetnames[sheet_index]]
        max_col = 40

        # ヘッダ行（"生徒ID" を含むセルが属する行）を検出
        header_row = None
        for r in range(1, 10):
            for c in range(1, 6):
                v = ws.cell(row=r, column=c).value
                if v and "生徒ID" in _normalize_label(v):
                    header_row = r
                    break
            if header_row:
                break
        if not header_row or header_row < 2:
            return {}
        group_row = header_row - 1

        # 行4（項目名）と行3（ブランド見出し）を収集
        item_labels = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=header_row, column=c).value
            lbl = _normalize_label(v)
            if lbl:
                item_labels[c] = lbl
        group_labels = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=group_row, column=c).value
            lbl = _normalize_label(v)
            if lbl:
                group_labels[c] = lbl

        # 隣接する単一文字見出しを結合（例: 白鳳 L3="ア" + M3="ン" → "アン"）
        merged_groups = {}
        sorted_gcols = sorted(group_labels.keys())
        i = 0
        while i < len(sorted_gcols):
            c = sorted_gcols[i]
            label = group_labels[c]
            j = i + 1
            # 隣接連続で、全部1文字なら結合
            while (j < len(sorted_gcols)
                   and sorted_gcols[j] == sorted_gcols[j - 1] + 1
                   and len(group_labels[sorted_gcols[j]]) == 1
                   and len(label) <= 3):
                label += group_labels[sorted_gcols[j]]
                j += 1
            merged_groups[c] = label
            i = j

        # 授業料位置 = 新ブランドの開始
        tuition_cols = sorted([c for c, lbl in item_labels.items() if lbl == "授業料"])

        # タイプA (二行ヘッダ) で授業料列が見つからない場合、
        # タイプB (単一行ヘッダ、列名に「ブランド+項目」を埋め込み) を試す
        if not tuition_cols:
            layout_b = _parse_type_b_header(item_labels, header_row)
            if layout_b:
                return layout_b
            return {}

        # 各授業料列に対応するブランド名を同定
        brand_at = {}
        used_group_cols = set()
        for tc in tuition_cols:
            # 候補: tc → tc+1 → tc-1 の順で merged_groups を探索
            label = None
            for offset in (0, 1, -1, 2):
                cand = tc + offset
                if cand in merged_groups and cand not in used_group_cols:
                    label = merged_groups[cand]
                    used_group_cols.add(cand)
                    break
            brand_at[tc] = label or f"Unknown_{get_column_letter(tc)}"

        # ブランドごとの列範囲（授業料列から次の授業料列 - 1 まで）
        TRAILING = {"入会金", "その他", "売上", "合計", "売上合計"}
        ranges = []
        for idx, start in enumerate(tuition_cols):
            brand = brand_at[start]
            if idx + 1 < len(tuition_cols):
                end = tuition_cols[idx + 1] - 1
            else:
                end = max_col
                for c in range(start, max_col + 1):
                    if c in item_labels and item_labels[c] in TRAILING:
                        end = c - 1
                        break
            ranges.append((start, end, brand))

        # 前置スタンドアロン列（生徒氏名より後、最初の授業料列より前）
        SKIP = {"No.", "生徒ID", "生徒氏名"}
        col_display = {}
        structured = {}  # letter -> (brand, item)

        first_brand_col = tuition_cols[0]
        for c in range(1, first_brand_col):
            if c in item_labels and item_labels[c] not in SKIP:
                letter = get_column_letter(c)
                item = item_labels[c]
                col_display[letter] = item
                structured[letter] = ("_leading", item)

        for start, end, brand in ranges:
            for c in range(start, end + 1):
                if c not in item_labels:
                    continue
                letter = get_column_letter(c)
                item = item_labels[c]
                col_display[letter] = f"{item}({brand})"
                structured[letter] = (brand, item)

        if ranges:
            last_end = ranges[-1][1]
            for c in range(last_end + 1, max_col + 1):
                if c in item_labels:
                    letter = get_column_letter(c)
                    col_display[letter] = item_labels[c]
                    structured[letter] = ("_trailing", item_labels[c])

        return {
            "col_display": col_display,
            "structured": structured,
            "header_row": header_row,
        }
    finally:
        wb.close()


# 正規レイアウト（小幡基準）: (brand, item) → canonical letter
_CANONICAL_MAP = {
    # 前置スタンドアロン
    ("_leading", "設備費"): "D",
    ("_leading", "空調費"): "D",
    # 学習塾
    ("学習塾", "授業料"): "E",
    ("学習塾", "月会費"): "F",
    ("学習塾", "テスト対策講習会費"): "G",
    ("学習塾", "模試代"): "H",
    ("学習塾", "講習会費"): "I",
    # プログラミング
    ("プログラミング", "授業料"): "J",
    ("プログラミング", "月会費"): "K",
    ("プログラミング", "講習会費"): "L",
    # アン(英会話)
    ("アン", "授業料"): "M",
    ("アン", "月会費"): "N",
    # そろばん
    ("そろばん", "授業料"): "O",
    ("そろばん", "月会費"): "P",
    ("そろばん", "講習会費"): "Q",
    # 筆っこ = アン美文字
    ("筆っこ", "授業料"): "R",
    ("筆っこ", "月会費"): "S",
    ("筆っこ", "講習会費"): "T",
    ("美文字", "授業料"): "R",
    ("美文字", "月会費"): "S",
    ("美文字", "講習会費"): "T",
    # 将棋
    ("将棋", "授業料"): "U",
    ("将棋", "月会費"): "V",
    ("将棋", "講習会費"): "W",
    # 後置スタンドアロン
    ("_trailing", "入会金"): "X",
    ("_trailing", "その他"): "Y",
    ("_trailing", "売上"): "Z",
    ("_trailing", "売上合計"): "Z",
}

# 読み込み時に適用される現在の列リマップ（src_letter -> canonical_letter）
# None なら恒等変換（既定の小幡レイアウト）
_CURRENT_REMAP: dict[str, str] | None = None


def build_canonical_remap(layout: dict) -> dict[str, str]:
    """
    detect_column_layout の結果から、校舎固有の列レターを
    正規レイアウト(小幡基準)のレターへマッピングする辞書を返す。
    学童など正規レイアウトに無いブランドの列は省かれる(= 削除)。
    """
    if not layout:
        return {}
    remap = {}
    structured = layout.get("structured", {})
    for src_letter, (brand, item) in structured.items():
        dst = _CANONICAL_MAP.get((brand, item))
        if dst:
            remap[src_letter] = dst
    return remap


def set_current_remap(remap: dict[str, str] | None) -> None:
    """read_excel_sales が次回以降適用する列リマップを設定する。"""
    global _CURRENT_REMAP
    _CURRENT_REMAP = remap if remap else None


def read_excel_sales(
    excel_path: str,
    sheet_index: int | None = 2,
) -> dict[str, dict]:
    """
    売上明細Excelを読み込む。
    sheet_index が None または負値なら自動検出する。
    指定したインデックスがレイアウト不一致でも自動検出にフォールバック。
    返り値: {生徒ID: {"name": 生徒名, "cols": {列レター: 金額}, "row": 行番号}}
    """
    if sheet_index is None or (isinstance(sheet_index, int) and sheet_index < 0):
        detected = detect_sales_sheet_index(excel_path)
        if detected is None:
            print(f"  警告: 売上明細シートを自動検出できませんでした: {excel_path}")
            return {}
        sheet_index = detected

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if sheet_index >= len(wb.sheetnames):
        print(f"  警告: シートindex {sheet_index} が存在しません: {excel_path}")
        wb.close()
        return {}

    ws = wb[wb.sheetnames[sheet_index]]

    # 指定シートがレイアウト的に売上明細でなければ自動検出にフォールバック
    if _score_sales_sheet(ws) < 5:
        wb.close()
        detected = detect_sales_sheet_index(excel_path)
        if detected is None or detected == sheet_index:
            print(f"  警告: 指定シートに生徒データが見つかりません (index={sheet_index}): {excel_path}")
            return {}
        print(f"  情報: index={sheet_index} は不一致、自動検出 index={detected} ({detected!r}) を使用")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[detected]]

    # ヘッダ行を動的検出してデータ開始行を決定
    header_row_detected = _find_header_row(ws)
    data_start = (header_row_detected + 1) if header_row_detected else 5

    result = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False), start=data_start):
        a_val = row[0].value
        if a_val is not None and isinstance(a_val, str) and "計" in a_val:
            break

        sid = row[1].value
        if sid is None:
            continue

        sid = str(int(sid)) if isinstance(sid, (int, float)) else str(sid)
        name = row[2].value or ""

        cols = {}
        for cell in row:
            if not hasattr(cell, "column_letter"):
                continue
            letter = cell.column_letter
            if letter in ("A", "B", "C"):
                continue
            val = parse_number(cell.value)
            if val != 0:
                cols[letter] = val

        # 正規レイアウトへの列リマップ適用（校舎別レイアウト差異を吸収）
        if _CURRENT_REMAP:
            remapped = {}
            for src_letter, v in cols.items():
                dst = _CURRENT_REMAP.get(src_letter)
                if dst:
                    remapped[dst] = remapped.get(dst, 0) + v
                # リマップに無い列は破棄（学童等、正規レイアウトに存在しないブランド）
            cols = remapped

        result[sid] = {"name": str(name), "cols": cols, "row": row_idx}

    wb.close()
    return result


# ====================================================================
# 照合チェック
# ====================================================================

def compare_student(
    excel_cols: dict[str, float],
    csv_cols: dict[str, float],
) -> list[tuple[str, float, float, float]]:
    """1人の生徒をExcel列ごとにCSVと比較 → [(列, Excel値, CSV値, 差額)]"""
    diffs = []
    all_cols = sorted(set(list(excel_cols.keys()) + list(csv_cols.keys())))

    for col in all_cols:
        if col == "Z":
            continue
        excel_val = excel_cols.get(col, 0)
        csv_val = csv_cols.get(col, 0)
        if isinstance(excel_val, str):
            continue
        if abs(excel_val - csv_val) >= 1:
            diffs.append((col, excel_val, csv_val, excel_val - csv_val))

    return diffs


def _find_nearest_class_info(
    year: int,
    month: int,
    class_info_files: dict[tuple[int, int], str],
) -> str | None:
    """対象月に最も近いクラス情報ファイルを返す"""
    if not class_info_files:
        return None
    target = year * 12 + month
    best_path = None
    best_dist = 9999
    for (cy, cm), path in class_info_files.items():
        dist = abs((cy * 12 + cm) - target)
        if dist < best_dist:
            best_dist = dist
            best_path = path
    return best_path


def _load_billing_by_priority(
    csv_paths_by_priority: list[str],
) -> dict[str, list[tuple[str, str, float]]]:
    """
    複数月のCSVを優先度順に読み込む。
    csv_paths_by_priority[0] = 対象月（最優先）
    csv_paths_by_priority[1:] = 前後月（対象月にない生徒のみ補完）

    各生徒の明細は1つの月からのみ取得（合算しない）。
    対象月にいれば対象月のデータ、いなければ前後月で見つかったデータを使う。
    """
    result: dict[str, list[tuple[str, str, float]]] = {}

    for path in csv_paths_by_priority:
        billing = read_billing_csv(path)
        for sid, entries in billing.items():
            if sid not in result:
                result[sid] = entries

    return result


def _get_adjacent_months(
    year: int, month: int, offset: int,
) -> list[tuple[int, int]]:
    """対象月の前後offset月を含むリストを返す"""
    months = []
    for delta in range(-offset, offset + 1):
        m = month + delta
        y = year
        while m < 1:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        months.append((y, m))
    return months


def _preload_all_billings(
    csv_paths_with_labels: list[tuple[str, str]],
) -> dict[str, dict[str, list[tuple[str, str, float]]]]:
    """
    全月のCSVを一括読み込み。
    返り値: {月ラベル: {生徒ID: [(ブランド名, カテゴリ名, 月額料金), ...]}}
    """
    all_billings: dict[str, dict[str, list[tuple[str, str, float]]]] = {}
    for path, label in csv_paths_with_labels:
        all_billings[label] = read_billing_csv(path)
    return all_billings


def _compute_monthly_billing(
    sid: str,
    all_billings: dict[str, dict[str, list[tuple[str, str, float]]]],
    month_col_labels: list[str],
    school_brands: dict[str, set[str]] | None,
) -> dict[str, dict[str, float]]:
    """
    事前読み込み済みのデータから生徒1人の各月別引落額を計算。
    返り値: {月ラベル: {列レター: 金額}}
    """
    monthly: dict[str, dict[str, float]] = {}

    for label in month_col_labels:
        billing = all_billings.get(label, {})
        entries = billing.get(sid, [])
        if school_brands is not None:
            student_brands = school_brands.get(sid)
            entries = filter_billing_by_school(entries, student_brands)
        agg = aggregate_csv_for_student(entries)
        monthly[label] = agg

    return monthly


def _find_similar_billing(
    sid: str,
    search_amount: float,
    all_billings: dict[str, dict[str, list[tuple[str, str, float]]]],
    month_col_labels: list[str],
    target_col: str = "",
) -> str:
    """
    差異がある項目について、前後月の全請求から
    同額または近い金額の請求を探してコメントを返す。
    target_col が指定された場合、そのExcel列にマッピングされるブランドの請求は除外する
    （既にマッピング済みのため）。
    """
    if search_amount == 0:
        return ""

    # target_colにマッピングされるブランドを特定（除外対象）
    exclude_brands = set()
    if target_col:
        for brand, (jcol, mcol) in BRAND_COLUMN_MAP.items():
            if jcol == target_col or mcol == target_col:
                exclude_brands.add(brand)

    hints = []
    seen = set()
    for label in month_col_labels:
        billing = all_billings.get(label, {})
        entries = billing.get(sid, [])
        for brand, category, amount in entries:
            if amount == 0:
                continue
            # 既にマッピング済みのブランドはスキップ
            if brand in exclude_brands:
                continue
            key = (label, brand, category, amount)
            if key in seen:
                continue
            if abs(amount - search_amount) < 1:
                hints.append(f"{label}「{brand}/{category}」{amount:,.0f}")
                seen.add(key)
            elif abs(search_amount) > 0 and abs(amount - search_amount) < abs(search_amount) * 0.3:
                hints.append(f"{label}「{brand}/{category}」{amount:,.0f}(近似)")
                seen.add(key)
    if hints:
        return " / ".join(hints[:3])
    return ""


# 講習会費系の列
KOSHUKAI_COLS = {"G", "I", "L", "Q", "T", "W"}
_DEFAULT_KOSHUKAI_COLS = set(KOSHUKAI_COLS)
_DEFAULT_COL_DISPLAY = None  # read_excel_sales 初回呼び出し前に埋める

# 定期費用の列（差額請求チェック対象）
RECURRING_COLS = {"D", "E", "F", "J", "K", "M", "N", "O", "P", "R", "S", "U", "V"}
_DEFAULT_RECURRING_COLS = set(RECURRING_COLS)


def _check_koshukai_alert(
    sid: str,
    col: str,
    excel_amount: float,
    all_billings: dict[str, dict[str, list[tuple[str, str, float]]]],
    month_col_labels: list[str],
    school_brands: dict[str, set[str]] | None,
) -> str:
    """
    講習会費アラート: 講習会費列で売上>0だがCSV=0の場合、
    全月の請求から同額を探す。見つかれば「○月に同額請求あり」、
    なければ「該当請求なし（要確認）」。
    """
    if col not in KOSHUKAI_COLS or excel_amount <= 0:
        return ""

    for label in month_col_labels:
        billing = all_billings.get(label, {})
        entries = billing.get(sid, [])
        if school_brands is not None:
            student_brands = school_brands.get(sid)
            if student_brands is not None:
                entries = filter_billing_by_school(entries, student_brands)
        agg = aggregate_csv_for_student(entries)
        if abs(agg.get(col, 0) - excel_amount) < 1:
            return f"⚡{label}に同額請求あり"

    return "⚠該当請求なし（要確認）"


def _detect_dropped_brands(
    sid: str,
    target_year: int,
    target_month: int,
    all_billings: dict[str, dict[str, list[tuple[str, str, float]]]],
    month_col_labels: list[str],
) -> dict[str, str]:
    """
    退会/コース変更検出: 前月にあったブランドが当月になくなった場合を検出。
    返り値: {列レター: アラートメッセージ}
    """
    target_label = f"{target_month}月"

    # 前月ラベルを特定
    prev_month = target_month - 1
    prev_year = target_year
    if prev_month < 1:
        prev_month = 12
        prev_year -= 1
    prev_label = f"{prev_month}月"

    if prev_label not in month_col_labels or target_label not in month_col_labels:
        return {}

    # 各月のブランド別金額を集計
    def get_brand_amounts(label):
        billing = all_billings.get(label, {})
        entries = billing.get(sid, [])
        brands = {}
        for brand, category, amount in entries:
            if not brand or category in ("設備費", "入会金", "0", "1",
                                          "入会時教材費", "入会時授業料1",
                                          "入会時授業料2", "入会時授業料3",
                                          "入会時授業料A", "入会時月会費",
                                          "入会時設備費", "家族割", "割引",
                                          "過不足金", "諸経費", "家賃",
                                          "総合指導管理費"):
                continue
            if brand not in brands:
                brands[brand] = 0
            brands[brand] += amount
        return {b: a for b, a in brands.items() if a > 0}

    prev_brands = get_brand_amounts(prev_label)
    curr_brands = get_brand_amounts(target_label)

    if not prev_brands:
        return {}

    dropped = set(prev_brands.keys()) - set(curr_brands.keys())
    if not dropped:
        return {}

    all_dropped = len(curr_brands) == 0

    alerts = {}
    for brand in dropped:
        # このブランドがマッピングされる列を特定
        if brand in BRAND_COLUMN_MAP:
            cols = list(BRAND_COLUMN_MAP[brand])
        else:
            cols = ["E", "F"]

        if all_dropped:
            msg = f"⚠{prev_month}月まで請求あり→{target_month}月から全コースなし（退会の可能性）"
        else:
            remaining = ", ".join(sorted(curr_brands.keys()))
            msg = f"⚠{prev_month}月まで{brand}請求あり→{target_month}月からなし（コース変更？残:{remaining}）"

        for col in cols:
            alerts[col] = msg

    return alerts


def _detect_amount_changes(
    sid: str,
    target_year: int,
    target_month: int,
    monthly_billing: dict[str, dict[str, float]],
    month_col_labels: list[str],
) -> dict[str, str]:
    """
    差額請求アラート: 定期費用が前月と当月で金額が変わった場合に警告。
    返り値: {列レター: アラートメッセージ}
    """
    target_label = f"{target_month}月"
    prev_month = target_month - 1
    if prev_month < 1:
        prev_month = 12
    prev_label = f"{prev_month}月"

    if prev_label not in month_col_labels or target_label not in month_col_labels:
        return {}

    prev_agg = monthly_billing.get(prev_label, {})
    curr_agg = monthly_billing.get(target_label, {})

    alerts = {}
    for col in RECURRING_COLS:
        pv = prev_agg.get(col, 0)
        cv = curr_agg.get(col, 0)
        if pv > 0 and cv > 0 and abs(pv - cv) >= 1:
            diff = cv - pv
            sign = "+" if diff > 0 else ""
            col_name = COL_DISPLAY.get(col, col)
            alerts[col] = f"💰前月{pv:,.0f}→当月{cv:,.0f}（{sign}{diff:,.0f} 差額請求の可能性）"

    return alerts


def _detect_amount_anomaly(
    sid: str,
    target_month: int,
    monthly_billing: dict[str, dict[str, float]],
    month_col_labels: list[str],
    excel_cols: dict[str, float],
) -> dict[str, str]:
    """
    金額異常検出: Excel=CSVで一致していても、他の月と比較して
    金額が大きく変動している場合に警告。
    前後月の金額と比較し、30%以上の乖離があればアラート。
    返り値: {列レター: アラートメッセージ}
    """
    target_label = f"{target_month}月"
    alerts = {}

    for col in RECURRING_COLS:
        curr_val = excel_cols.get(col, 0)
        if not isinstance(curr_val, (int, float)) or curr_val <= 0:
            continue

        # 他月の同じ列の金額を収集
        other_vals = []
        for label in month_col_labels:
            if label == target_label:
                continue
            agg = monthly_billing.get(label, {})
            v = agg.get(col, 0)
            if v > 0:
                other_vals.append((label, v))

        if not other_vals:
            continue

        # 他月の代表値（最頻値に近いもの）
        amounts = [v for _, v in other_vals]
        median = sorted(amounts)[len(amounts) // 2]

        # 30%以上乖離かつ1000円以上の差
        if median > 0 and abs(curr_val - median) >= max(median * 0.3, 1000):
            diff = curr_val - median
            sign = "+" if diff > 0 else ""
            ref_months = " / ".join(f"{l}:{v:,.0f}" for l, v in other_vals[:3])
            alerts[col] = f"⚠金額変動注意: 当月{curr_val:,.0f} vs 他月({ref_months})（{sign}{diff:,.0f}）"

    return alerts


def _get_y_col_details(
    sid: str,
    all_billings: dict[str, dict[str, list[tuple[str, str, float]]]],
    month_col_labels: list[str],
) -> str:
    """Y列(その他)にマッピングされるCSV明細の内訳を返す"""
    details = []
    for label in month_col_labels:
        billing = all_billings.get(label, {})
        entries = billing.get(sid, [])
        for brand, category, amount in entries:
            if amount == 0:
                continue
            mapped = _map_to_column(brand, category)
            if mapped == "Y":
                details.append(f"{label}「{brand}/{category}」{amount:,.0f}")
    if details:
        return " / ".join(details[:5])
    return ""


def run_check(
    school_name: str,
    month_label: str,
    csv_paths: list[str],
    csv_paths_with_labels: list[tuple[str, str]],
    excel_path: str,
    sheet_index: int = 2,
    school_brands: dict[str, set[str]] | None = None,
    withdrawn_sids: set[str] | None = None,
    pair_year: int = 0,
    pair_month: int = 0,
    all_class_sids: set[str] | None = None,
    mid_month_withdrawals: dict[str, str] | None = None,
) -> list[CheckResult]:
    """
    1校舎・1ヶ月分のチェックを実行。
    csv_paths: 対象月±2ヶ月分のCSVパスリスト（優先度順）
    csv_paths_with_labels: [(path, "11月"), (path, "12月"), ...] 表示順
    school_brands: {生徒ID: {この校舎で受講しているブランド}} or None
    """
    csv_names = [os.path.basename(p) for p in csv_paths]
    print(f"\n{'='*80}")
    print(f"  [{school_name}] {month_label} 照合チェック")
    print(f"  CSV:   {csv_names[0]} 他{len(csv_paths)-1}ヶ月分" if len(csv_paths) > 1
          else f"  CSV:   {csv_names[0]}")
    print(f"  Excel: {os.path.basename(excel_path)}")
    if school_brands is not None:
        print(f"  校舎フィルタ: 有効（クラス情報から{len(school_brands)}生徒分）")
    print(f"{'='*80}")

    billing = _load_billing_by_priority(csv_paths)
    sales = read_excel_sales(excel_path, sheet_index)
    grades = read_grades_from_csv(csv_paths[0])  # 対象月CSVから学年取得

    month_col_labels = [label for _, label in csv_paths_with_labels]

    # 月別引落額計算用: 全CSVを事前に一括読み込み
    all_billings = _preload_all_billings(csv_paths_with_labels)

    ok_count = 0
    col_only_count = 0
    total_diff_count = 0
    not_in_csv_count = 0
    no_billing_count = 0
    withdrawn_count = 0
    not_enrolled_count = 0
    results: list[CheckResult] = []

    _withdrawn = withdrawn_sids or set()
    _mid_wd = mid_month_withdrawals or {}

    for sid, sdata in sorted(sales.items(), key=lambda x: x[1]["row"]):
        name = sdata["name"]
        excel_cols = sdata["cols"]
        row_num = sdata["row"]

        # 在籍不明（クラス情報に存在しない生徒に売上がある）
        excel_total_raw = excel_cols.get("Z", 0)
        if isinstance(excel_total_raw, str):
            excel_total_raw = 0
        if all_class_sids is not None and sid not in all_class_sids and excel_total_raw > 0:
            not_enrolled_count += 1
            results.append(CheckResult(
                school=school_name,
                month_label=month_label,
                result_type="NOT_ENROLLED",
                sid=sid, name=name, row=row_num,
                excel_total=excel_total_raw,
                month_columns=month_col_labels,
                grade=grades.get(sid, ""),
                remarks={"Z": "クラス情報に在籍なし（休会・退会・未登録の可能性）"},
            ))
            continue

        # 退会者なのに売上がある
        if sid in _withdrawn and excel_total_raw > 0:
            withdrawn_count += 1
            results.append(CheckResult(
                school=school_name,
                month_label=month_label,
                result_type="WITHDRAWN",
                sid=sid,
                name=name,
                row=row_num,
                excel_total=excel_total_raw,
                month_columns=month_col_labels,
                grade=grades.get(sid, ""),
            ))
            continue

        if sid not in billing:
            total = excel_cols.get("Z", 0)
            if total != 0 and not isinstance(total, str):
                not_in_csv_count += 1
                results.append(CheckResult(
                    school=school_name,
                    month_label=month_label,
                    result_type="NOT_IN_CSV",
                    sid=sid,
                    name=name,
                    row=row_num,
                    excel_total=total,
                    month_columns=month_col_labels,
                ))
            continue

        entries = billing[sid]
        if school_brands is not None:
            student_brands = school_brands.get(sid)
            entries = filter_billing_by_school(entries, student_brands)
        csv_agg = aggregate_csv_for_student(entries)
        diffs = compare_student(excel_cols, csv_agg)

        if diffs:
            excel_total = excel_cols.get("Z", 0)
            if isinstance(excel_total, str):
                excel_total = 0
            csv_total = sum(csv_agg.values())
            total_match = abs(excel_total - csv_total) < 1

            # 各月別の引落額を計算（事前読み込み済みデータを使用）
            monthly = _compute_monthly_billing(
                sid, all_billings, month_col_labels, school_brands,
            )

            # 退会/コース変更検出（請求ベース）
            dropped_alerts = _detect_dropped_brands(
                sid, pair_year, pair_month,
                all_billings, month_col_labels,
            )
            # 差額請求アラート
            amount_change_alerts = _detect_amount_changes(
                sid, pair_year, pair_month,
                monthly, month_col_labels,
            )
            # 当月退会アラート（ブランド単位・日割り/回数調整の可能性）
            mid_wd_col_alerts = {}
            if sid in _mid_wd:
                brand_dates = _mid_wd[sid]  # {ブランド名: 退会日}
                for wd_brand, wd_date in brand_dates.items():
                    # このブランドがマッピングされる列を特定
                    if wd_brand in BRAND_COLUMN_MAP:
                        cols = list(BRAND_COLUMN_MAP[wd_brand])
                    else:
                        cols = ["E", "F"]
                    # 設備費(D)も退会なら0にすべき
                    cols.append("D")
                    msg = f"⚠{wd_brand} {wd_date}退会（金額調整の可能性）"
                    for c in cols:
                        if c not in mid_wd_col_alerts:
                            mid_wd_col_alerts[c] = msg

            # 項目レベルで「売上あり請求なし」と「その他の差異」を分離
            # 売上>0 かつ 請求額と不一致 → 売上あり請求なし（Y列は対象外）
            unbilled_diffs = [
                d for d in diffs
                if d[1] > 0 and d[0] != "Y"
            ]
            other_diffs = [
                d for d in diffs
                if not (d[1] > 0 and d[0] != "Y")
            ]

            if unbilled_diffs:
                no_billing_count += 1
                remarks = {}
                for col, ev, cv, diff in unbilled_diffs:
                    hints = []
                    if col in mid_wd_col_alerts:
                        hints.append(mid_wd_col_alerts[col])
                    # 講習会費アラート
                    koshu = _check_koshukai_alert(
                        sid, col, ev, all_billings,
                        month_col_labels, school_brands,
                    )
                    if koshu:
                        hints.append(koshu)
                    # 退会/コース変更アラート
                    if col in dropped_alerts:
                        hints.append(dropped_alerts[col])
                    # 差額請求アラート
                    if col in amount_change_alerts:
                        hints.append(amount_change_alerts[col])
                    # 類似請求検索（講習会費以外）
                    if col not in KOSHUKAI_COLS:
                        similar = _find_similar_billing(
                            sid, ev, all_billings, month_col_labels, col,
                        )
                        if similar:
                            hints.append(similar)
                    if hints:
                        remarks[col] = " / ".join(hints)
                results.append(CheckResult(
                    school=school_name,
                    month_label=month_label,
                    result_type="NO_BILLING",
                    sid=sid, name=name, row=row_num,
                    diffs=unbilled_diffs,
                    excel_total=excel_total,
                    csv_total=csv_total,
                    monthly_billing=monthly,
                    month_columns=month_col_labels,
                    remarks=remarks,
                    grade=grades.get(sid, ""),
                ))

            if other_diffs:
                if total_match:
                    col_only_count += 1
                    rtype = "TOTAL_MATCH"
                else:
                    total_diff_count += 1
                    rtype = "TOTAL_DIFF"
                other_remarks = {}
                for col, ev, cv, diff in other_diffs:
                    hints = []
                    if col in mid_wd_col_alerts:
                        hints.append(mid_wd_col_alerts[col])
                    if col in dropped_alerts:
                        hints.append(dropped_alerts[col])
                    if col in amount_change_alerts:
                        hints.append(amount_change_alerts[col])
                    # Y列(その他)でCSV>0、Excel=0の場合、CSVの内訳を表示
                    if col == "Y" and ev == 0 and cv > 0:
                        y_details = _get_y_col_details(
                            sid, all_billings, month_col_labels,
                        )
                        if y_details:
                            hints.append(f"⚠売上未計上: {y_details}")
                    if ev != 0 or cv != 0:
                        search_amt = ev if ev != 0 else cv
                        similar = _find_similar_billing(
                            sid, search_amt, all_billings, month_col_labels, col,
                        )
                        if similar:
                            hints.append(similar)
                    if hints:
                        other_remarks[col] = " / ".join(hints)
                results.append(CheckResult(
                    school=school_name,
                    month_label=month_label,
                    result_type=rtype,
                    sid=sid, name=name, row=row_num,
                    diffs=other_diffs,
                    excel_total=excel_total,
                    csv_total=csv_total,
                    monthly_billing=monthly,
                    month_columns=month_col_labels,
                    remarks=other_remarks,
                    grade=grades.get(sid, ""),
                ))
            elif not unbilled_diffs:
                # unbilledもotherもない（ありえないが安全策）
                pass
        else:
            # 差異なしでも当月退会ブランドがあればアラート
            if sid in _mid_wd and excel_total_raw > 0:
                brand_dates = _mid_wd[sid]
                wd_remarks = {}
                wd_diffs = []
                for wd_brand, wd_date in brand_dates.items():
                    if wd_brand in BRAND_COLUMN_MAP:
                        cols = list(BRAND_COLUMN_MAP[wd_brand])
                    else:
                        cols = ["E", "F"]
                    cols.append("D")
                    msg = f"⚠{wd_brand} {wd_date}退会（金額調整の可能性）"
                    for c in cols:
                        ev = excel_cols.get(c, 0)
                        if isinstance(ev, (int, float)) and ev > 0:
                            cv = csv_agg.get(c, 0)
                            wd_diffs.append((c, ev, cv, ev - cv))
                            wd_remarks[c] = msg
                if wd_diffs:
                    monthly = _compute_monthly_billing(
                        sid, all_billings, month_col_labels, school_brands,
                    )
                    results.append(CheckResult(
                        school=school_name,
                        month_label=month_label,
                        result_type="NO_BILLING",
                        sid=sid, name=name, row=row_num,
                        diffs=wd_diffs,
                        excel_total=excel_total_raw,
                        csv_total=sum(csv_agg.values()),
                        monthly_billing=monthly,
                        month_columns=month_col_labels,
                        remarks=wd_remarks,
                        grade=grades.get(sid, ""),
                    ))
                    no_billing_count += 1
                else:
                    ok_count += 1
            else:
                # 金額異常検出（Excel=CSVでも他月と大きく乖離）
                monthly = _compute_monthly_billing(
                    sid, all_billings, month_col_labels, school_brands,
                )
                anomaly_alerts = _detect_amount_anomaly(
                    sid, pair_month, monthly, month_col_labels, excel_cols,
                )
                if anomaly_alerts and excel_total_raw > 0:
                    anomaly_diffs = []
                    for col, msg in anomaly_alerts.items():
                        ev = excel_cols.get(col, 0)
                        cv = csv_agg.get(col, 0)
                        if isinstance(ev, (int, float)) and ev > 0:
                            anomaly_diffs.append((col, ev, cv, ev - cv))
                    if anomaly_diffs:
                        no_billing_count += 1
                        results.append(CheckResult(
                            school=school_name,
                            month_label=month_label,
                            result_type="NO_BILLING",
                            sid=sid, name=name, row=row_num,
                            diffs=anomaly_diffs,
                            excel_total=excel_total_raw,
                            csv_total=sum(csv_agg.values()),
                            monthly_billing=monthly,
                            month_columns=month_col_labels,
                            remarks=anomaly_alerts,
                            grade=grades.get(sid, ""),
                        ))
                    else:
                        ok_count += 1
                else:
                    ok_count += 1

    # コンソール出力
    print(f"\n■ 結果サマリー")
    print(f"  Excel生徒数:       {len(sales)}")
    print(f"  完全一致 (OK):     {ok_count}")
    print(f"  列配分違い(合計一致): {col_only_count}")
    print(f"  合計不一致(要確認):  {total_diff_count}")
    print(f"  売上あり請求なし:    {no_billing_count}")
    print(f"  退会者売上あり:     {withdrawn_count}")
    print(f"  月謝未計上:        {not_in_csv_count}")

    _print_details(results)

    return results


def _print_details(results: list[CheckResult]) -> None:
    """差異詳細をコンソール出力"""
    critical = [r for r in results if r.result_type == "TOTAL_DIFF"]
    not_in_csv = [r for r in results if r.result_type == "NOT_IN_CSV"]
    col_only = [r for r in results if r.result_type == "TOTAL_MATCH"]

    if critical:
        print(f"\n{'='*80}")
        print(f"  ★★★ 合計不一致（要確認）: {len(critical)}件 ★★★")
        print(f"{'='*80}")
        for r in critical:
            tdiff = r.excel_total - r.csv_total
            sign = "+" if tdiff > 0 else ""
            print(f"\n  生徒ID={r.sid} {r.name} (Row {r.row})"
                  f"  合計差額={sign}{tdiff:,.0f}")
            for col, ev, cv, diff in r.diffs:
                disp = COL_DISPLAY.get(col, col)
                s = "+" if diff > 0 else ""
                print(f"    {disp:20s}  Excel={ev:>10,.0f}  "
                      f"CSV={cv:>10,.0f}  差={s}{diff:,.0f}")
            print(f"    {'─'*60}")
            print(f"    {'合計':20s}  Excel={r.excel_total:>10,.0f}  "
                  f"CSV={r.csv_total:>10,.0f}  差={sign}{tdiff:,.0f}")

    if not_in_csv:
        print(f"\n{'─'*80}")
        print(f"  月謝未計上: {len(not_in_csv)}件")
        print(f"{'─'*80}")
        for r in not_in_csv:
            print(f"  生徒ID={r.sid} {r.name} "
                  f"(Row {r.row}) Excel売上={r.excel_total:,.0f}")

    if col_only:
        print(f"\n{'─'*80}")
        print(f"  列配分違い（合計は一致）: {len(col_only)}件")
        print(f"{'─'*80}")
        for r in col_only:
            print(f"  生徒ID={r.sid} {r.name} (Row {r.row})")
            for col, ev, cv, diff in r.diffs:
                disp = COL_DISPLAY.get(col, col)
                s = "+" if diff > 0 else ""
                print(f"    {disp:20s}  Excel={ev:>10,.0f}  "
                      f"CSV={cv:>10,.0f}  差={s}{diff:,.0f}")


# ====================================================================
# CSV出力
# ====================================================================

def write_results_csv(
    output_path: str,
    all_results: list[CheckResult],
) -> None:
    """照合結果をCSVに出力（月別引落額付き）"""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 全結果から月カラムのユニオンを取得
    all_month_cols: list[str] = []
    for r in all_results:
        for mc in r.month_columns:
            if mc not in all_month_cols:
                all_month_cols.append(mc)
    all_month_cols.sort()

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        header = [
            "校舎", "月", "重要度", "種別", "生徒ID", "生徒名",
            "行番号", "項目", "売上",
        ]
        for mc in all_month_cols:
            header.append(f"{mc}引落")
        header.extend(["差額", "合計差額"])
        writer.writerow(header)

        for r in all_results:
            if r.result_type == "NOT_IN_CSV":
                row_data = [
                    r.school, r.month_label, "★要確認", "月謝未計上",
                    r.sid, r.name, r.row, "売上合計",
                    r.excel_total,
                ]
                row_data.extend([0] * len(all_month_cols))
                row_data.extend([r.excel_total, r.excel_total])
                writer.writerow(row_data)
            elif r.diffs:
                total_diff = r.excel_total - r.csv_total
                total_match = abs(total_diff) < 1

                # 売上あり・請求なし判定
                if r.csv_total == 0 and r.excel_total > 0:
                    severity = "★売上あり請求なし"
                    kind = "請求漏れ"
                elif total_match:
                    severity = "列配分違い"
                    kind = "合計一致"
                else:
                    severity = "★要確認"
                    kind = "合計不一致"

                for col, ev, cv, diff in r.diffs:
                    disp = COL_DISPLAY.get(col, col)
                    row_data = [
                        r.school, r.month_label, severity, kind,
                        r.sid, r.name, r.row, disp, ev,
                    ]
                    for mc in all_month_cols:
                        monthly_agg = r.monthly_billing.get(mc, {})
                        row_data.append(monthly_agg.get(col, 0))
                    row_data.extend([diff, total_diff])
                    writer.writerow(row_data)


# ====================================================================
# メイン
# ====================================================================

def main():
    config = load_config()

    csv_base_dir = config["csv_base_dir"]
    class_info_dir = config.get("class_info_dir", "")
    output_dir = config.get("output_dir", "C:/Users/USER/Documents/照合結果")
    schools = []
    for s in config["schools"]:
        s = dict(s)  # コピーして元configを壊さない
        kw = s.pop("school_keywords", [])
        schools.append(SchoolConfig(**s, school_keywords=tuple(kw)))

    print("=" * 80)
    print("  売上明細 vs 請求データ 照合チェック（全校舎対応）")
    print(f"  実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  対象校舎: {', '.join(s.name for s in schools)}")
    print(f"  チェック範囲: 対象月 ±2ヶ月")
    print("=" * 80)

    # 請求CSV自動検出
    csv_files = discover_csv_files(csv_base_dir)
    print(f"\n  検出された請求CSV: {len(csv_files)}ヶ月分")
    for (y, m), path in sorted(csv_files.items()):
        print(f"    {y}年{m}月 ← {os.path.basename(path)}")

    # クラス情報自動検出
    class_info_files = {}
    if class_info_dir:
        class_info_files = discover_class_info_files(class_info_dir)
        if class_info_files:
            print(f"\n  検出されたクラス情報: {len(class_info_files)}ヶ月分")
            for (y, m), path in sorted(class_info_files.items()):
                print(f"    {y}年{m}月 ← {os.path.basename(path)}")
        else:
            print(f"\n  ⚠ クラス情報が見つかりません: {class_info_dir}")

    all_results: list[CheckResult] = []

    for school in schools:
        print(f"\n{'#'*80}")
        print(f"  校舎: {school.name}")
        print(f"  Excel: {school.excel_dir}")
        print(f"{'#'*80}")

        excel_files = discover_excel_files(school)
        if not excel_files:
            print(f"  ⚠ Excelファイルが見つかりません: {school.excel_dir}")
            continue

        print(f"  検出されたExcel: {len(excel_files)}ヶ月分")
        for (y, m), path in sorted(excel_files.items()):
            print(f"    {y}年{m}月 ← {os.path.basename(path)}")

        pairs = match_months(csv_files, excel_files)
        if not pairs:
            print(f"  ⚠ CSVとExcelで一致する月がありません")
            continue

        print(f"  マッチした月: {len(pairs)}ヶ月")

        for pair in pairs:
            # 対象月を先頭に、±2ヶ月分のCSVパスを優先度順に収集
            csv_paths = [pair.csv_path]  # 対象月が最優先
            adjacent = _get_adjacent_months(pair.year, pair.month, 2)
            for ym in adjacent:
                if ym in csv_files and csv_files[ym] != pair.csv_path:
                    csv_paths.append(csv_files[ym])

            # 表示用: 月ラベル付きCSVリスト（時系列順）
            csv_paths_with_labels = []
            for ym in adjacent:
                if ym in csv_files:
                    csv_paths_with_labels.append(
                        (csv_files[ym], f"{ym[0]}年{ym[1]}月")
                    )

            # クラス情報から校舎別ブランドフィルタ＋退会者リストを構築
            school_brands = None
            withdrawn_sids = None
            all_class_sids = None
            mid_month_wd = {}
            if school.school_keywords and class_info_files:
                best_ci = _find_nearest_class_info(
                    pair.year, pair.month, class_info_files,
                )
                if best_ci:
                    school_brands = read_class_info(
                        best_ci, school.school_keywords,
                    )
                    withdrawn_sids = read_withdrawn_students(
                        best_ci, school.school_keywords,
                        pair.year, pair.month,
                    )
                    all_class_sids = read_all_class_sids(
                        best_ci, school.school_keywords,
                    )
                    mid_month_wd = read_mid_month_withdrawals(
                        best_ci, school.school_keywords,
                        pair.year, pair.month,
                    )

            results = run_check(
                school_name=school.name,
                month_label=pair.label,
                csv_paths=csv_paths,
                csv_paths_with_labels=csv_paths_with_labels,
                excel_path=pair.excel_path,
                sheet_index=school.sheet_index,
                school_brands=school_brands,
                withdrawn_sids=withdrawn_sids,
                pair_year=pair.year,
                pair_month=pair.month,
                all_class_sids=all_class_sids,
                mid_month_withdrawals=mid_month_wd,
            )
            all_results.extend(results)

    # 結果CSV出力
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"照合結果_{timestamp}.csv")
    write_results_csv(output_path, all_results)

    # 全校舎サマリー
    print(f"\n{'='*80}")
    print(f"  全校舎サマリー")
    print(f"{'='*80}")
    school_names = sorted(set(r.school for r in all_results))
    for sn in school_names:
        school_results = [r for r in all_results if r.school == sn]
        critical = sum(1 for r in school_results if r.result_type == "TOTAL_DIFF")
        col_only = sum(1 for r in school_results if r.result_type == "TOTAL_MATCH")
        not_csv = sum(1 for r in school_results if r.result_type == "NOT_IN_CSV")
        print(f"  {sn}: 合計不一致={critical}件, "
              f"列配分違い={col_only}件, 月謝未計上={not_csv}件")

    print(f"\n■ 照合結果CSVを出力しました: {output_path}")
    print()


if __name__ == "__main__":
    main()
